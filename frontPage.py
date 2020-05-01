import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, colors, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

import pickle

BOLDFONT = Font(bold=True)

COLORS = {'red':'ff0000','orange':'ffbf00','yellow':'ffff00','green':'a6db94'}

FULLBORDER = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

def shortenCols(wb):
	ws = wb['OVERVIEW']
	listCols = ['C','D','E','F','G','H','I','J',
				'K','L','M','N','O','P','Q','R',
				'S','T','U','V','W','X']
	for col in listCols:
		ws.column_dimensions[col].width = 2.15
	ws.column_dimensions['B'].width=10

	ws.sheet_view.showGridLines = False

def getColor(response):
	### takes in a response and returns what color the square should be
	# print(response[2])
	if "No Response" in response[0]:
		return 'orange'
	rDict = {
		'green':['4:','5:','1:','9Z:','2C:','3U:','2E:'],
		'red':['3T:','3P:','3N:','6A:','3M:','3D:','3C:','3F:','3B:'],
		'yellow':['2B:','3H:','2D:','8:','2A:']
	}

	for color in rDict:
		for code in rDict[color]:
			if code in response[2]:
				return color


def createFrontPage(responses,wb,titleDict):
	wb.create_sheet('OVERVIEW',0)
	ws = wb['OVERVIEW']

	shortenCols(wb)

	### ws.cell(row=y,column=x,value=z)
	cRow = 2

	for workID in responses:
		ws.cell(row=cRow,column=1,value=workID)
		ws.cell(row=cRow,column=1).font = BOLDFONT
		ws.cell(row=cRow-1,column=3,value=titleDict[workID])
		ws.cell(row=cRow-1,column=3).font = BOLDFONT
		for ticket in responses[workID]:
			ws.cell(row=cRow,column=2,value=ticket)
			ws.cell(row=cRow,column=2).number_format = "@"
			cCol = 3
			for response in responses[workID][ticket]:
				color = getColor(response)
				color = COLORS[color]
				ws.cell(row=cRow,column=cCol).fill = PatternFill(start_color=color,end_color=color,fill_type='solid')
				ws.cell(row=cRow,column=cCol).border = FULLBORDER
				cCol += 1
			cRow += 1
		cRow += 2