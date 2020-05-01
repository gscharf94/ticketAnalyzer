import openpyxl

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, colors, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

def returnCoordString(col,row):
	dicty = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H'}

	column = dicty[col]
	row = str(row)

	return column+row

def colorCell(col,row,color,ws):
	colors = {'red':'ff0000','orange':'ffbf00','yellow':'ffff00','green':'a6db94'}

	coordString = returnCoordString(col,row)
	ws[coordString].fill = PatternFill(start_color=colors[color],end_color=colors[color],fill_type='solid')


def createSpreadsheet(workOrders,saveName,titleDict):
	wb = openpyxl.Workbook()
	wb2 = openpyxl.Workbook()

	wsDicty = {}
	# print(titleDict)
	wsDict2 = {}

	workOrders1 = {}
	workOrders2 = {}

	for workOrder in workOrders:
		if 'TMS' in titleDict[workOrder]:
			wsDict2[workOrder] = wb2.create_sheet(workOrder)
			workOrders2[workOrder] = workOrders[workOrder]
		else:
			wsDicty[workOrder] = wb.create_sheet(workOrder)
			workOrders1[workOrder] = workOrders[workOrder]

	for sheet in wsDicty:
		wsDicty[sheet].column_dimensions['A'].width = 10
		wsDicty[sheet].column_dimensions['B'].width = 22
		wsDicty[sheet].column_dimensions['C'].width = 41.43
		wsDicty[sheet].column_dimensions['D'].width = 51.57
		wsDicty[sheet].column_dimensions['E'].width = 40

	for sheet in wsDict2:
		wsDict2[sheet].column_dimensions['A'].width = 10
		wsDict2[sheet].column_dimensions['B'].width = 22
		wsDict2[sheet].column_dimensions['C'].width = 41.43
		wsDict2[sheet].column_dimensions['D'].width = 51.57
		wsDict2[sheet].column_dimensions['E'].width = 40

	for workOrder in workOrders1:
		enterSheet(workOrder,wsDicty,workOrders1)

	for workOrder in workOrders2:
		enterSheet(workOrder,wsDict2,workOrders2)

	for sheet in wsDicty:
		colorAppropriateCells(wsDicty[sheet])

	for sheet in wsDict2:
		colorAppropriateCells(wsDict2[sheet])


	firstPage = wb['Sheet']
	wb.remove(firstPage)

	firstPage2 = wb2['Sheet']
	wb2.remove(firstPage2)

	wb.save(saveName)
	saveName2 = saveName[:-5]+" TMS.xlsx"
	wb2.save(saveName2)

	return [saveName,saveName2]



def enterSheet(workOrder,wsDicty,workOrders):
	ws = wsDicty[workOrder]

	sX = 1
	sY = 1

	thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

	for ticket in workOrders[workOrder]:
		sX = 1
		ws.cell(row=sY,column=sX,value=ticket)
		sY += 1
		sX += 1
		for response in workOrders[workOrder][ticket]:
			for cell in response:
				ws.cell(row=sY,column=sX,value=cell)
				ws[returnCoordString(sX,sY)].alignment = Alignment(wrap_text=True,horizontal='left',vertical='top')
				ws[returnCoordString(sX,sY)].border = thin_border
				sX += 1
			sY += 1
			sX = 2

def colorAppropriateCells(ws):
	col = 1
	row = 1

	for y in range(250):
		for x in range(4):
			value = ws.cell(row=y+1,column=x+1).value
			# print(value)
			if value == None:
				pass
			else:
				if "4:" in value or "5:" in value or "1:" in value or "9Z:" in value or "2C:" in value or "3U:" in value or "2E:" in value:
					colorCell(x+1,y+1,'green',ws)
					colorCell(x-1,y+1,'green',ws)
					colorCell(x+2,y+1,'green',ws)
					colorCell(x,y+1,'green',ws)
				elif "No Response" in value:
					colorCell(x+1,y+1,'orange',ws)
					colorCell(x+2,y+1,'orange',ws)
					colorCell(x+3,y+1,'orange',ws)
					colorCell(x+4,y+1,'orange',ws)
				elif "3T:" in value or "3P:" in value or "3N:" in value or "6A:" in value or "3M:" in value or "3D:" in value or "3C:" in value or "3F:" in value or "3B:" in value:
					colorCell(x+1,y+1,'red',ws)
					colorCell(x-1,y+1,'red',ws)
					colorCell(x+2,y+1,'red',ws)
					colorCell(x,y+1,'red',ws)
				elif "2B:" in value or "3H:" in value or "2D:" in value or "8:" in value or "2A:" in value or "2A:" in value:
					colorCell(x+1,y+1,'yellow',ws)
					colorCell(x-1,y+1,'yellow',ws)
					colorCell(x+2,y+1,'yellow',ws)
					colorCell(x,y+1,'yellow',ws)
	for y in range(250):
		value = ws.cell(row=y+1,column=5).value
		# print(value)
		if value == None:
			pass
		else:
			if "Respondant:" in value:
				ws.cell(row=y+1,column=5,value="")

def addTitles(titleDict,sysPath):
	wb = load_workbook(sysPath)

	for workOrder in titleDict:
		for sheet in wb.worksheets:
			if workOrder == sheet.title:
				wb[workOrder].insert_rows(1)
				wb[workOrder]['B1'] = titleDict[workOrder]
				wb[workOrder].merge_cells('B1:E1')
				wb[workOrder]['B1'].font = Font(bold=True,size='20')
				wb[workOrder]['B1'].alignment = Alignment(horizontal='center')

	wb.save(sysPath)


